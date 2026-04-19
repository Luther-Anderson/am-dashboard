#!/usr/bin/env python3
"""
parse_excel.py — Extract AM KPIs from the weekly Google Sheet XLSX download.
Writes data.json (Cash-In, Meetings from Excel; pipeline_snapshot preserved).

Usage:  python3 parse_excel.py <path-to-xlsx>
"""

import sys
import json
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils.datetime import from_excel as xl_to_dt
except ImportError:
    sys.exit("❌  openpyxl not installed. Run: pip3 install openpyxl")

# ─── Config ────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent

# Excel name → (owner_id, full_name, fin_row, ops_meetings_row, ops_avg_time_row, ops_week_pct_row, ops_qtd_row)
# fin_row           : row in 'Sales Financial KPIs'   (total cash-in per AM)
# ops_meetings_row  : row in 'Sales Operational KPIs' (# of meetings)
# ops_avg_time_row  : row in 'Sales Operational KPIs' (avg meeting time in minutes)
# ops_week_pct_row  : row in 'Sales Operational KPIs' (last-week attainment %)
# ops_qtd_row       : row in 'Sales Operational KPIs' (QTD attainment %)
AM_MAP = {
    'Edouard':  ('751897671', 'Edouard Daenen',    19, 75, 78, 76, 77),
    'Viviana':  ('90532029',  'Viviana El Mouak',  24, 82, 85, 83, 84),
    'Patricia': ('76991649',  'Patricia Chalmeta', 29, 89, 92, 90, 91),
    'Antolin':  ('51221997',  'Antolin Lera Jeuk', 34, 96, 99, 97, 98),
}

# Team-level attainment rows (Viviana excluded from target in sheet formula)
TEAM_MTG_WEEK_PCT_ROW = 69
TEAM_MTG_QTD_ROW      = 70

COL_START = 2   # Column B (first week)
COL_END   = 14  # Column N (last week — 13 weeks total)

# ─── Helpers ───────────────────────────────────────────────────────────────
def to_monday(d):
    return (d - timedelta(days=d.weekday())).strftime('%Y-%m-%d')

def cell_to_date(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)) and val > 1:
        try:
            return xl_to_dt(val).strftime('%Y-%m-%d')
        except Exception:
            pass
    if isinstance(val, str):
        try:
            return datetime.fromisoformat(val).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None

def quarter_mondays(quarter_start, quarter_end):
    s = datetime.fromisoformat(quarter_start)
    e = datetime.fromisoformat(quarter_end)
    days_ahead = (7 - s.weekday()) % 7
    start_mon = s + timedelta(days=days_ahead)
    weeks, cur = [], start_mon
    while cur.date() <= e.date():
        weeks.append(cur.strftime('%Y-%m-%d'))
        cur += timedelta(weeks=1)
    return weeks

def last_value(ws, row_num):
    """Return the last non-null numeric value in cols B–N (most recent completed week)."""
    result = None
    for col in range(COL_START, COL_END + 1):
        v = ws.cell(row=row_num, column=col).value
        if isinstance(v, (int, float)):
            result = v
    return result

def read_row(ws, row_num):
    vals = []
    for col in range(COL_START, COL_END + 1):
        v = ws.cell(row=row_num, column=col).value
        try:
            vals.append(float(v) if v not in (None, '') else 0.0)
        except (TypeError, ValueError):
            vals.append(0.0)
    return vals

# ─── Core parsing ──────────────────────────────────────────────────────────
def parse(xlsx_path, manual):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    fin_ws = wb['Sales Financial KPIs']
    ops_ws = wb['Sales Operational KPIs']

    valid_weeks = set(quarter_mondays(manual['quarter_start'], manual['quarter_end']))

    # Resolve week dates from header row 1, cols B–N
    sheet_weeks = []
    for col in range(COL_START, COL_END + 1):
        raw = fin_ws.cell(row=1, column=col).value
        iso = cell_to_date(raw)
        if iso:
            sheet_weeks.append(to_monday(datetime.fromisoformat(iso)))
        else:
            sheet_weeks.append(None)

    resolved_in_quarter = [w for w in sheet_weeks if w and w in valid_weeks]
    if not resolved_in_quarter:
        print('  ⚠️  Header dates outside current quarter (stale cache). Computing from quarter_start.')
        computed = quarter_mondays(manual['quarter_start'], manual['quarter_end'])
        sheet_weeks = (computed + [None] * 13)[:13]

    resolved = [w for w in sheet_weeks if w]
    print(f'  📅  Weeks resolved: {resolved[0]} → {resolved[-1]}  ({len(resolved)} weeks)')

    # Team-level meeting attainment (Viviana excluded from target by sheet formula)
    meetings_attainment = {
        'team': {
            'week_pct': last_value(ops_ws, TEAM_MTG_WEEK_PCT_ROW),
            'qtd_pct':  last_value(ops_ws, TEAM_MTG_QTD_ROW),
        }
    }

    closed_won, meetings_out, avg_meeting_time_out = [], [], []

    for am_key, (owner_id, owner_name, fin_row, ops_mtg_row, ops_amt_row, ops_week_pct_row, ops_qtd_row) in AM_MAP.items():
        cash_vals = read_row(fin_ws, fin_row)
        meet_vals = read_row(ops_ws, ops_mtg_row)
        amt_vals  = read_row(ops_ws, ops_amt_row)
        meetings_attainment[owner_id] = {
            'week_pct': last_value(ops_ws, ops_week_pct_row),
            'qtd_pct':  last_value(ops_ws, ops_qtd_row),
        }

        for i, week in enumerate(sheet_weeks):
            if week is None or week not in valid_weeks:
                continue

            cash  = cash_vals[i]
            meets = meet_vals[i]
            amt   = amt_vals[i]

            if cash > 0:
                closed_won.append({
                    'week': week, 'owner_id': owner_id, 'owner_name': owner_name,
                    'amount': round(cash, 2), 'count': 1,
                })
            if meets > 0:
                meetings_out.append({
                    'week': week, 'owner_id': owner_id, 'owner_name': owner_name,
                    'count': int(meets),
                })
            if amt > 0:
                avg_meeting_time_out.append({
                    'week': week, 'owner_id': owner_id, 'owner_name': owner_name,
                    'avg_minutes': round(amt, 1),
                })

    return closed_won, meetings_out, avg_meeting_time_out, meetings_attainment

# ─── Main ──────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        sys.exit('Usage: python3 parse_excel.py <path-to-xlsx>')

    xlsx_path = Path(sys.argv[1]).expanduser().resolve()
    if not xlsx_path.exists():
        sys.exit(f'❌  File not found: {xlsx_path}')

    print(f'\n📊  Parsing {xlsx_path.name} …')

    manual = json.loads((SCRIPT_DIR / 'manual-data.json').read_text())
    closed_won, meetings, avg_meeting_time, meetings_attainment = parse(xlsx_path, manual)

    print(f'  ✅  Cash-In rows:        {len(closed_won)}')
    print(f'  ✅  Meeting rows:        {len(meetings)}')
    print(f'  ✅  Avg meeting time:    {len(avg_meeting_time)}')

    # Preserve pipeline_snapshot and closed_lost from HubSpot (updated separately)
    data_path = SCRIPT_DIR / 'data.json'
    pipeline_snapshot = []
    closed_lost = []
    if data_path.exists():
        existing = json.loads(data_path.read_text())
        pipeline_snapshot = existing.get('pipeline_snapshot', [])
        closed_lost       = existing.get('closed_lost', [])
        print(f'  ✅  Pipeline rows:      {len(pipeline_snapshot)} (kept from last HubSpot sync)')
        print(f'  ✅  Closed Lost rows:   {len(closed_lost)} (kept from last HubSpot sync)')

    output = {
        'updated_at':        datetime.now(timezone.utc).isoformat(),
        'quarter':           manual['quarter'],
        'quarter_start':     manual['quarter_start'],
        'closed_won':           sorted(closed_won,        key=lambda r: (r['week'], r['owner_id'])),
        'closed_lost':          sorted(closed_lost,       key=lambda r: (r['week'], r['owner_id'])),
        'meetings':             sorted(meetings,          key=lambda r: (r['week'], r['owner_id'])),
        'avg_meeting_time':     sorted(avg_meeting_time, key=lambda r: (r['week'], r['owner_id'])),
        'meetings_attainment':  meetings_attainment,
        'pipeline_snapshot':    pipeline_snapshot,
    }

    data_path.write_text(json.dumps(output, indent=2))
    print(f'\n✅  data.json written\n')

if __name__ == '__main__':
    main()
